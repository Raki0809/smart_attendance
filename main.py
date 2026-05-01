import cv2
import os
import numpy as np
import pandas as pd
from datetime import datetime
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ====== REGISTERED STUDENTS ======
registered_students = ["Rakesh", "Arun", "Priya", "Kumar", "Anita"]

dataset_path = "dataset"
face_cascade = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")

faces = []
labels = []
label_map = {}
current_label = 0

# ===== Load Dataset =====
for person_name in os.listdir(dataset_path):
    person_folder = os.path.join(dataset_path, person_name)
    if os.path.isdir(person_folder):
        label_map[current_label] = person_name
        for image_name in os.listdir(person_folder):
            img_path = os.path.join(person_folder, image_name)
            img = cv2.imread(img_path, cv2.IMREAD_GRAYSCALE)
            img = cv2.resize(img, (200, 200))
            faces.append(img)
            labels.append(current_label)
        current_label += 1

labels = np.array(labels)

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.train(faces, labels)

print("Press SPACE to start attendance")

cap = cv2.VideoCapture(0)

attendance_started = False
start_time = None
attendance_file = "attendance.xlsx"

# Create Excel file
if not os.path.exists(attendance_file):
    df = pd.DataFrame(columns=["Name", "Date", "Time", "Status"])
    df.to_excel(attendance_file, index=False)

marked_students = {}

while True:
    ret, frame = cap.read()
    if not ret:
        break

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
    faces_detected = face_cascade.detectMultiScale(gray, 1.3, 5)

    current_time = time.time()

    if attendance_started:
        elapsed = current_time - start_time

        for (x, y, w, h) in faces_detected:
            face = gray[y:y+h, x:x+w]
            face = cv2.resize(face, (200, 200))
            label, confidence = recognizer.predict(face)

            if confidence < 80:
                name = label_map[label]

                if name not in marked_students:
                    now = datetime.now()
                    date = now.strftime("%Y-%m-%d")
                    time_now = now.strftime("%H:%M:%S")

                    if elapsed <= 1800:
                        status = "On Time"
                        color = "90EE90"   # Green
                    elif elapsed <= 2100:
                        status = "Late"
                        color = "FFFF99"   # Yellow
                    else:
                        continue

                    df = pd.read_excel(attendance_file)
                    new_entry = pd.DataFrame([[name, date, time_now, status]],
                                             columns=["Name", "Date", "Time", "Status"])
                    df = pd.concat([df, new_entry], ignore_index=True)
                    df.to_excel(attendance_file, index=False)

                    # Apply Color
                    wb = load_workbook(attendance_file)
                    ws = wb.active
                    last_row = ws.max_row
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for col in range(1, 5):
                        ws.cell(row=last_row, column=col).fill = fill
                    wb.save(attendance_file)

                    marked_students[name] = status

                cv2.putText(frame, name, (x, y-10),
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0,255,0), 2)

            cv2.rectangle(frame, (x,y), (x+w,y+h), (255,0,0), 2)

        # After 35 mins mark absent
        if elapsed > 2100:
            for student in registered_students:
                if student not in marked_students:
                    df = pd.read_excel(attendance_file)
                    now = datetime.now()
                    new_entry = pd.DataFrame([[student,
                                               now.strftime("%Y-%m-%d"),
                                               "--",
                                               "Absent"]],
                                             columns=["Name", "Date", "Time", "Status"])
                    df = pd.concat([df, new_entry], ignore_index=True)
                    df.to_excel(attendance_file, index=False)

                    wb = load_workbook(attendance_file)
                    ws = wb.active
                    last_row = ws.max_row
                    fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                    for col in range(1, 5):
                        ws.cell(row=last_row, column=col).fill = fill
                    wb.save(attendance_file)

            print("Attendance Completed")
            break

    cv2.imshow("Smart Attendance System", frame)

    key = cv2.waitKey(1)

    if key == 32 and not attendance_started:
        attendance_started = True
        start_time = time.time()
        print("Attendance Started")

    if key == 27:
        break

cap.release()
cv2.destroyAllWindows()