import cv2
import pickle
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

today = datetime.now().strftime("%Y-%m-%d")
attendance_file = f"{today}_attendance.xlsx"

dataset_path = "dataset"

if not os.path.exists(dataset_path):
    print("Dataset folder not found")
    exit()

all_students = []
for name in os.listdir(dataset_path):
    if os.path.isdir(os.path.join(dataset_path, name)):
        all_students.append(name)

if os.path.exists(attendance_file):
    wb = load_workbook(attendance_file)
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Name","Date","Time","Status","Confidence"])
    wb.save(attendance_file)

ws = wb.active

if not os.path.exists("model.yml"):
    print("Model not found")
    exit()

recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("model.yml")

with open("labels.pickle","rb") as f:
    label_map = pickle.load(f)

face_cascade = cv2.CascadeClassifier("face_model.xml")

if face_cascade.empty():
    print("Error loading face_model.xml")
    exit()

cap = cv2.VideoCapture(0)

marked_names = set()
start_time = datetime.now()

print("Attendance system started")

while True:
    ret, frame = cap.read()

    if not ret:
        print("Camera error")
        break

    frame = cv2.resize(frame, (640, 480))
    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    faces = face_cascade.detectMultiScale(gray, 1.3, 5)

    for (x,y,w,h) in faces:
        face = gray[y:y+h, x:x+w]
        face = cv2.resize(face,(200,200))

        label, confidence = recognizer.predict(face)

        if confidence < 70:
            name = label_map[label]
        else:
            name = "Unknown"

        if name != "Unknown" and name not in marked_names:
            now = datetime.now()

            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")

            minutes_passed = (now - start_time).total_seconds() / 60

            if minutes_passed <= 30:
                status = "Present"
            elif minutes_passed <= 35:
                status = "Late"
            else:
                status = "Absent"

            ws.append([name, date_str, time_str, status, round(confidence,2)])
            marked_names.add(name)

            print(name, status)

    session_minutes = (datetime.now() - start_time).total_seconds() / 60

    if session_minutes > 35:
        print("Attendance session ended")
        break

    if cv2.waitKey(1) == 27:
        break

cap.release()
cv2.destroyAllWindows()

date_now = datetime.now().strftime("%Y-%m-%d")

for student in all_students:
    if student not in marked_names:
        ws.append([student, date_now, "--", "Absent", "--"])

green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
yellow = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
red = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")

for row in ws.iter_rows(min_row=2, max_col=5):
    status = row[3].value

    if status == "Present":
        color = green
    elif status == "Late":
        color = yellow
    else:
        color = red

    for cell in row:
        cell.fill = color

wb.save(attendance_file)